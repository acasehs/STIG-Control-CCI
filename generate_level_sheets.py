#!/usr/bin/env python3
"""
STIG Control Level Reference Sheet Generator

This script processes NIST 800-53 controls organized by Defense Levels (DL-1 through DL-6)
and generates an Excel workbook with:
- Individual sheets for each level with control details and CCI mappings
- A summary sheet with charts and tables broken out by control family

Usage:
    python generate_level_sheets.py [--input INPUT_FILE] [--output OUTPUT_FILE]

The input file can be:
- A JSON file with level_data structure
- A CSV file with columns: DL-1, DL-2, DL-3, DL-4, DL-5, DL-6
"""

import json
import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

# tkinter for file dialog (built into Python)
try:
    import tkinter as tk
    from tkinter import filedialog
    HAS_TKINTER = True
except ImportError:
    HAS_TKINTER = False

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Required packages not found. Installing...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'pandas', 'openpyxl'])
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter


def open_file_dialog() -> str:
    """Open a file dialog to select an input file. Returns the selected file path or None."""
    if not HAS_TKINTER:
        print("Warning: tkinter not available for file dialog. Use --input to specify file.")
        return None

    # Create and hide the root window
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)  # Bring dialog to front

    # Open file dialog
    file_path = filedialog.askopenfilename(
        title="Select Level Data File",
        filetypes=[
            ("All Supported", "*.xlsx *.xls *.json *.csv"),
            ("Excel Files", "*.xlsx *.xls"),
            ("JSON Files", "*.json"),
            ("CSV Files", "*.csv"),
            ("All Files", "*.*")
        ]
    )

    root.destroy()

    return file_path if file_path else None


# Default level data based on the provided spreadsheet
# Format: Each level contains a list of control identifiers
DEFAULT_LEVEL_DATA = {
    "DL-1 DODIN": [
        "AT-01", "AT-02", "AT-02(01)", "AT-02(02)", "CM-10(01)"
    ],
    "DL-2 MCEN": [
        "AC-04", "AC-04(01)", "AC-04(02)", "AC-04(03)", "AC-04(04)"
    ],
    "DL-3 MITSC/IPN/ISN/Data Center": [
        "AC-19(04)", "AC-20(02)", "AC-23", "AP-01", "AP-02"
    ],
    "DL-4": [
        "PE-02", "PE-02(01)", "PE-02(02)", "PE-02(03)", "PE-03"
    ],
    "DL-5 System HW/SW/OS": [
        "AC-06(08)", "AC-06(10)", "AC-07", "AC-07(02)", "AC-08"
    ],
    "DL-6 Application": [
        "AC-01", "AC-02", "AC-02(01)", "AC-02(02)", "AC-02(03)"
    ]
}


def normalize_control_id(control_id: str) -> str:
    """
    Normalize control identifier to double-digit format.
    Examples:
        AC-1 -> AC-01
        AC-2(1) -> AC-02(01)
        AT-1 -> AT-01
    """
    if not control_id:
        return ""

    # Pattern to match control IDs like AC-1, AC-01, AC-2(1), AC-02(01)
    pattern = r'^([A-Z]{2})-(\d+)(?:\((\d+)\))?$'
    match = re.match(pattern, control_id.strip().upper())

    if match:
        family = match.group(1)
        control_num = int(match.group(2))
        enhancement = match.group(3)

        if enhancement:
            return f"{family}-{control_num:02d}({int(enhancement):02d})"
        else:
            return f"{family}-{control_num:02d}"

    return control_id.strip().upper()


def load_controls_data(filepath: str) -> dict:
    """Load controls data from JSON file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Create a lookup dictionary by normalized control identifier
    controls_lookup = {}
    for control in data:
        control_id = normalize_control_id(control.get('Control Identifier', ''))
        if control_id:
            controls_lookup[control_id] = {
                'name': control.get('Control (or Control Enhancement) Name', ''),
                'text': control.get('Control Text', ''),
                'discussion': control.get('Discussion', ''),
                'related_controls': control.get('Related Controls', '')
            }

    return controls_lookup


def load_comparison_data(script_dir: Path) -> dict:
    """Load Rev 4 to Rev 5 comparison data if available."""
    comparison_path = script_dir / 'r4_r5_comparison.json'
    if comparison_path.exists():
        with open(comparison_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return {
            'withdrawn': set(data.get('withdrawn_rev4_only', [])),
            'new': set(data.get('new_rev5_only', []))
        }
    return {'withdrawn': set(), 'new': set()}


def load_cci_data(filepath: str) -> dict:
    """Load CCI mappings from JSON file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Create a lookup dictionary: control_id -> list of CCIs
    cci_lookup = defaultdict(list)
    for item in data:
        control_id = normalize_control_id(item.get('Control', ''))
        if control_id:
            cci_lookup[control_id].append({
                'cci_number': item.get('CCI Number', ''),
                'description': item.get('Description', ''),
                'index': item.get('Index', '')
            })

    return dict(cci_lookup)


def get_control_family(control_id: str) -> str:
    """Extract control family from control identifier."""
    if not control_id or not control_id.strip():
        return "Unknown"
    match = re.match(r'^([A-Z]{2,3})-', control_id.strip().upper())
    return match.group(1) if match else "Unknown"


def validate_control_id(control_id: str) -> bool:
    """Check if a string looks like a valid control ID."""
    if not control_id or not control_id.strip():
        return False
    # Valid patterns: AC-01, AC-01(01), AC-1, etc.
    pattern = r'^[A-Z]{2,3}-\d+(\(\d+\))?$'
    return bool(re.match(pattern, control_id.strip().upper()))


def get_family_name(family_code: str) -> str:
    """Get full family name from family code."""
    family_names = {
        # NIST 800-53 Rev 5 Families
        'AC': 'Access Control',
        'AT': 'Awareness and Training',
        'AU': 'Audit and Accountability',
        'CA': 'Assessment, Authorization, and Monitoring',
        'CM': 'Configuration Management',
        'CP': 'Contingency Planning',
        'IA': 'Identification and Authentication',
        'IR': 'Incident Response',
        'MA': 'Maintenance',
        'MP': 'Media Protection',
        'PE': 'Physical and Environmental Protection',
        'PL': 'Planning',
        'PM': 'Program Management',
        'PS': 'Personnel Security',
        'PT': 'PII Processing and Transparency',
        'RA': 'Risk Assessment',
        'SA': 'System and Services Acquisition',
        'SC': 'System and Communications Protection',
        'SI': 'System and Information Integrity',
        'SR': 'Supply Chain Risk Management',
        # Privacy Controls (Appendix J)
        'AP': 'Authority and Purpose',
        'AR': 'Accountability, Audit, and Risk Management',
        'DI': 'Data Quality and Integrity',
        'DM': 'Data Minimization and Retention',
        'IP': 'Individual Participation and Redress',
        'SE': 'Security',
        'TR': 'Transparency',
        'UL': 'Use Limitation',
        # Other
        'Unknown': 'Unknown/Invalid'
    }
    return family_names.get(family_code, family_code)


def load_level_data_from_csv(filepath: str) -> dict:
    """Load level data from CSV file."""
    df = pd.read_csv(filepath)
    level_data = {}

    for col in df.columns:
        controls = df[col].dropna().tolist()
        normalized_controls = [normalize_control_id(c) for c in controls if c]
        level_data[col] = [c for c in normalized_controls if c]

    return level_data


def load_level_data_from_json(filepath: str) -> dict:
    """Load level data from JSON file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Normalize all control IDs
    level_data = {}
    for level, controls in data.items():
        level_data[level] = [normalize_control_id(c) for c in controls if c]

    return level_data


def load_level_data_from_excel(filepath: str, sheet_name: str = None) -> dict:
    """
    Load level data from Excel file (.xlsx or .xls).

    Expected format: Columns are level names, rows contain control IDs.
    Example:
        DL-1 DODIN | DL-2 MCEN | DL-3 MITSC...
        AT-01      | AC-04     | AC-19(04)
        AT-02      | AC-04(01) | AC-20(02)
        ...        | ...       | ...

    Args:
        filepath: Path to Excel file
        sheet_name: Optional sheet name to read (defaults to first sheet)
    """
    # Read the Excel file
    if sheet_name:
        df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)
    else:
        df = pd.read_excel(filepath, dtype=str)

    level_data = {}
    invalid_entries = []

    for col in df.columns:
        # Get all non-null values from the column
        controls = df[col].dropna().tolist()
        valid_controls = []

        for c in controls:
            if not c or not str(c).strip():
                continue
            normalized = normalize_control_id(str(c).strip())
            if normalized:
                if validate_control_id(normalized):
                    valid_controls.append(normalized)
                else:
                    # Still include it but warn
                    invalid_entries.append((col, str(c).strip(), normalized))
                    valid_controls.append(normalized)

        level_data[col] = valid_controls

    # Warn about potentially invalid entries
    if invalid_entries:
        print(f"\nWarning: {len(invalid_entries)} entries don't match standard control ID format:")
        for col, original, normalized in invalid_entries[:10]:  # Show first 10
            print(f"  [{col}] '{original}' -> '{normalized}'")
        if len(invalid_entries) > 10:
            print(f"  ... and {len(invalid_entries) - 10} more")
        print()

    return level_data


def create_level_sheet(wb: Workbook, level_name: str, controls: list,
                       controls_lookup: dict, cci_lookup: dict) -> dict:
    """
    Create a worksheet for a specific level with control details and CCI mappings.
    Returns statistics for summary.
    """
    # Create safe sheet name (max 31 chars)
    safe_name = level_name[:31].replace('/', '-').replace('\\', '-')
    ws = wb.create_sheet(title=safe_name)

    # Define styles
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['Control ID', 'Control Name', 'Control Text', 'CCI Numbers', 'CCI Count', 'Family']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Statistics tracking
    stats = {
        'total_controls': 0,
        'total_ccis': 0,
        'families': defaultdict(int),
        'family_ccis': defaultdict(int),
        'unknown_controls': [],  # Track controls with Unknown family
        'not_in_reference': []   # Track controls not found in reference data
    }

    # Populate data
    row = 2
    for control_id in controls:
        normalized_id = normalize_control_id(control_id)
        if not normalized_id:
            continue

        control_info = controls_lookup.get(normalized_id, {})
        ccis = cci_lookup.get(normalized_id, [])
        family = get_control_family(normalized_id)

        # Join CCI numbers
        cci_numbers = ', '.join([c['cci_number'] for c in ccis]) if ccis else 'N/A'
        cci_count = len(ccis)

        # Track problematic entries
        if family == "Unknown":
            stats['unknown_controls'].append(normalized_id)
        if not control_info:
            stats['not_in_reference'].append(normalized_id)

        # Update stats
        stats['total_controls'] += 1
        stats['total_ccis'] += cci_count
        stats['families'][family] += 1
        stats['family_ccis'][family] += cci_count

        # Write row
        ws.cell(row=row, column=1, value=normalized_id).border = border
        ws.cell(row=row, column=2, value=control_info.get('name', 'N/A')).border = border

        text_cell = ws.cell(row=row, column=3, value=control_info.get('text', 'N/A')[:1000])
        text_cell.alignment = Alignment(wrap_text=True)
        text_cell.border = border

        cci_cell = ws.cell(row=row, column=4, value=cci_numbers)
        cci_cell.alignment = Alignment(wrap_text=True)
        cci_cell.border = border

        ws.cell(row=row, column=5, value=cci_count).border = border
        ws.cell(row=row, column=6, value=family).border = border

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10

    # Freeze header row
    ws.freeze_panes = 'A2'

    return stats


def create_summary_sheet(wb: Workbook, all_stats: dict, level_names: list):
    """Create summary sheet with charts and tables."""
    ws = wb.create_sheet(title="Summary", index=0)

    # Styles
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="42A5F5", end_color="42A5F5", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.cell(row=1, column=1, value="STIG Control Level Summary Report")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.merge_cells('A1:G1')

    # Overview Table
    ws.cell(row=3, column=1, value="Level Overview")
    ws.cell(row=3, column=1).font = Font(bold=True, size=14)

    overview_headers = ['Level', 'Total Controls', 'Total CCIs', 'Avg CCIs/Control']
    for col, header in enumerate(overview_headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    row = 5
    for level_name in level_names:
        stats = all_stats.get(level_name, {})
        total_controls = stats.get('total_controls', 0)
        total_ccis = stats.get('total_ccis', 0)
        avg_ccis = round(total_ccis / total_controls, 2) if total_controls > 0 else 0

        ws.cell(row=row, column=1, value=level_name[:30]).border = border
        ws.cell(row=row, column=2, value=total_controls).border = border
        ws.cell(row=row, column=3, value=total_ccis).border = border
        ws.cell(row=row, column=4, value=avg_ccis).border = border
        row += 1

    # Create bar chart for controls per level
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Controls per Level"
    chart1.y_axis.title = "Count"
    chart1.x_axis.title = "Level"

    data = Reference(ws, min_col=2, min_row=4, max_row=row-1, max_col=2)
    cats = Reference(ws, min_col=1, min_row=5, max_row=row-1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    chart1.width = 15
    chart1.height = 10
    ws.add_chart(chart1, "F3")

    # Family Breakdown Table
    family_row_start = row + 2
    ws.cell(row=family_row_start, column=1, value="Controls by Family Across Levels")
    ws.cell(row=family_row_start, column=1).font = Font(bold=True, size=14)

    # Collect all families
    all_families = set()
    for stats in all_stats.values():
        all_families.update(stats.get('families', {}).keys())
    all_families = sorted(all_families)

    # Family table headers
    family_headers = ['Family', 'Family Name'] + [l[:15] for l in level_names] + ['Total']
    header_row = family_row_start + 1
    for col, header in enumerate(family_headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    # Family data
    data_row = header_row + 1
    family_totals = defaultdict(int)

    for family in all_families:
        ws.cell(row=data_row, column=1, value=family).border = border
        ws.cell(row=data_row, column=2, value=get_family_name(family)).border = border

        row_total = 0
        for col, level_name in enumerate(level_names, 3):
            stats = all_stats.get(level_name, {})
            count = stats.get('families', {}).get(family, 0)
            ws.cell(row=data_row, column=col, value=count).border = border
            row_total += count
            family_totals[family] += count

        ws.cell(row=data_row, column=len(level_names) + 3, value=row_total).border = border
        data_row += 1

    # Create stacked bar chart for families by level
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "stacked"
    chart2.style = 10
    chart2.title = "Control Families by Level"
    chart2.y_axis.title = "Controls"

    # Data for chart (families as series, levels as categories)
    for i, family in enumerate(all_families):
        family_data_row = header_row + 1 + i
        data = Reference(ws, min_col=3, min_row=family_data_row, max_col=2 + len(level_names))
        chart2.add_data(data, titles_from_data=False)
        if chart2.series:
            chart2.series[-1].tx = SeriesLabel(v=family)

    cats = Reference(ws, min_col=3, min_row=header_row, max_col=2 + len(level_names))
    chart2.set_categories(cats)
    chart2.width = 18
    chart2.height = 12
    ws.add_chart(chart2, "F" + str(family_row_start))

    # CCI Coverage by Family Table
    cci_row_start = data_row + 2
    ws.cell(row=cci_row_start, column=1, value="CCI Count by Family Across Levels")
    ws.cell(row=cci_row_start, column=1).font = Font(bold=True, size=14)

    cci_header_row = cci_row_start + 1
    for col, header in enumerate(family_headers, 1):
        cell = ws.cell(row=cci_header_row, column=col, value=header)
        cell.fill = subheader_fill
        cell.font = header_font
        cell.border = border

    cci_data_row = cci_header_row + 1
    for family in all_families:
        ws.cell(row=cci_data_row, column=1, value=family).border = border
        ws.cell(row=cci_data_row, column=2, value=get_family_name(family)).border = border

        row_total = 0
        for col, level_name in enumerate(level_names, 3):
            stats = all_stats.get(level_name, {})
            count = stats.get('family_ccis', {}).get(family, 0)
            ws.cell(row=cci_data_row, column=col, value=count).border = border
            row_total += count

        ws.cell(row=cci_data_row, column=len(level_names) + 3, value=row_total).border = border
        cci_data_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 45
    for i in range(len(level_names)):
        ws.column_dimensions[get_column_letter(i + 3)].width = 18

    return ws


def create_cci_detail_sheet(wb: Workbook, level_name: str, controls: list,
                           controls_lookup: dict, cci_lookup: dict):
    """Create a detailed CCI breakdown sheet for a level."""
    safe_name = (level_name[:25] + " CCIs").replace('/', '-').replace('\\', '-')
    ws = wb.create_sheet(title=safe_name)

    # Styles
    header_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['Control ID', 'Control Name', 'CCI Number', 'CCI Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    row = 2
    for control_id in controls:
        normalized_id = normalize_control_id(control_id)
        if not normalized_id:
            continue

        control_info = controls_lookup.get(normalized_id, {})
        ccis = cci_lookup.get(normalized_id, [])

        if not ccis:
            # Still show control even if no CCIs
            ws.cell(row=row, column=1, value=normalized_id).border = border
            ws.cell(row=row, column=2, value=control_info.get('name', 'N/A')).border = border
            ws.cell(row=row, column=3, value='N/A').border = border
            ws.cell(row=row, column=4, value='No CCIs mapped').border = border
            row += 1
        else:
            for cci in ccis:
                ws.cell(row=row, column=1, value=normalized_id).border = border
                ws.cell(row=row, column=2, value=control_info.get('name', 'N/A')).border = border
                ws.cell(row=row, column=3, value=cci['cci_number']).border = border
                desc_cell = ws.cell(row=row, column=4, value=cci['description'][:500])
                desc_cell.alignment = Alignment(wrap_text=True)
                desc_cell.border = border
                row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 80

    ws.freeze_panes = 'A2'


def create_rev4_only_sheet(wb: Workbook, rev4_controls: dict,
                           r4_controls_lookup: dict, r4_cci_lookup: dict):
    """
    Create a separate sheet for Rev 4-only (withdrawn) controls.

    Args:
        wb: Workbook to add sheet to
        rev4_controls: Dict of {level_name: [control_ids]} for Rev 4-only controls
        r4_controls_lookup: Rev 4 controls data
        r4_cci_lookup: Rev 4 CCI mappings
    """
    ws = wb.create_sheet(title="Rev 4 Only (Withdrawn)")

    # Styles
    header_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")  # Orange
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['Level', 'Control ID', 'Control Name', 'Control Text', 'CCI Numbers', 'CCI Count', 'Note']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    row = 2
    total_controls = 0

    for level_name, controls in rev4_controls.items():
        for control_id in controls:
            normalized_id = normalize_control_id(control_id)
            if not normalized_id:
                continue

            control_info = r4_controls_lookup.get(normalized_id, {})
            ccis = r4_cci_lookup.get(normalized_id, [])

            cci_numbers = ', '.join([c['cci_number'] for c in ccis]) if ccis else 'N/A'
            cci_count = len(ccis)

            ws.cell(row=row, column=1, value=level_name[:25]).border = border
            ws.cell(row=row, column=2, value=normalized_id).border = border
            ws.cell(row=row, column=3, value=control_info.get('name', 'N/A')).border = border

            text_cell = ws.cell(row=row, column=4, value=control_info.get('text', 'N/A')[:500])
            text_cell.alignment = Alignment(wrap_text=True)
            text_cell.border = border

            cci_cell = ws.cell(row=row, column=5, value=cci_numbers)
            cci_cell.alignment = Alignment(wrap_text=True)
            cci_cell.border = border

            ws.cell(row=row, column=6, value=cci_count).border = border
            ws.cell(row=row, column=7, value="Withdrawn in Rev 5").border = border

            row += 1
            total_controls += 1

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 18

    ws.freeze_panes = 'A2'

    return total_controls


def main():
    parser = argparse.ArgumentParser(
        description='Generate STIG Control Level Reference Sheets with CCI Mappings'
    )
    parser.add_argument(
        '--input', '-i',
        help='Input file (JSON, CSV, or Excel) with level data. Uses default data if not specified.'
    )
    parser.add_argument(
        '--sheet', '-s',
        help='Sheet name to read from Excel file (defaults to first sheet)'
    )
    parser.add_argument(
        '--output', '-o',
        default=None,
        help='Output Excel file path (default: <input_name>_CCI_Breakdown.xlsx)'
    )
    parser.add_argument(
        '--controls', '-c',
        default=None,
        help='Path to controls JSON file (auto-detects Rev 5, falls back to Rev 4)'
    )
    parser.add_argument(
        '--cci', '-cci',
        default=None,
        help='Path to CCI mappings JSON file (auto-detects Rev 5, falls back to Rev 4)'
    )
    parser.add_argument(
        '--detailed-cci',
        action='store_true',
        help='Generate detailed CCI sheets for each level'
    )
    parser.add_argument(
        '--no-gui',
        action='store_true',
        help='Skip file dialog and use --input or default data (for CLI/scripted usage)'
    )

    args = parser.parse_args()

    # Get script directory for relative paths
    script_dir = Path(__file__).parent

    # Determine input file (from argument, file dialog, or default)
    input_file = args.input

    # Open file dialog by default unless --input is provided or --no-gui is set
    if not input_file and not args.no_gui:
        print("Opening file selection dialog...")
        input_file = open_file_dialog()
        if not input_file:
            print("No file selected. Using default level data.")

    # Load level data
    if input_file:
        input_path = Path(input_file)
        suffix = input_path.suffix.lower()
        if suffix == '.csv':
            level_data = load_level_data_from_csv(str(input_path))
            print(f"Loaded level data from CSV: {input_file}")
        elif suffix in ['.xlsx', '.xls']:
            level_data = load_level_data_from_excel(str(input_path), args.sheet)
            sheet_info = f" (sheet: {args.sheet})" if args.sheet else " (first sheet)"
            print(f"Loaded level data from Excel: {input_file}{sheet_info}")
        else:
            level_data = load_level_data_from_json(str(input_path))
            print(f"Loaded level data from JSON: {input_file}")
    else:
        level_data = DEFAULT_LEVEL_DATA
        print("Using default level data")

    # Determine output file name (auto-generate from input if not specified)
    if args.output:
        output_file = args.output
    elif input_file:
        # Name output based on input file: "myfile.xlsx" -> "myfile_CCI_Breakdown.xlsx"
        input_path = Path(input_file)
        output_file = str(input_path.parent / f"{input_path.stem}_CCI_Breakdown.xlsx")
    else:
        # Default when using embedded data
        output_file = "STIG_Control_Level_Reference.xlsx"

    # Load controls and CCI data with Rev 5 -> Rev 4 fallback
    def find_data_file(user_path, rev5_name, rev4_name):
        """Find data file: use user path if provided, else try Rev 5, then Rev 4."""
        if user_path:
            path = Path(user_path) if Path(user_path).is_absolute() else script_dir / user_path
            if path.exists():
                return path, None
            else:
                raise FileNotFoundError(f"Specified file not found: {path}")

        # Try Rev 5 first
        rev5_path = script_dir / rev5_name
        if rev5_path.exists():
            return rev5_path, "Rev 5"

        # Fall back to Rev 4
        rev4_path = script_dir / rev4_name
        if rev4_path.exists():
            return rev4_path, "Rev 4 (fallback)"

        raise FileNotFoundError(f"No data files found. Looked for {rev5_name} and {rev4_name}")

    controls_path, controls_rev = find_data_file(args.controls, 'r5controls.json', 'r4controls.json')
    cci_path, cci_rev = find_data_file(args.cci, 'rev5cci.json', 'rev4cci.json')

    rev_info = controls_rev or "custom"
    print(f"Loading controls from {controls_path} ({rev_info})...")
    controls_lookup = load_controls_data(str(controls_path))
    print(f"Loaded {len(controls_lookup)} controls")

    rev_info = cci_rev or "custom"
    print(f"Loading CCI mappings from {cci_path} ({rev_info})...")
    cci_lookup = load_cci_data(str(cci_path))
    print(f"Loaded CCIs for {len(cci_lookup)} controls")

    # Load Rev 4 to Rev 5 comparison data
    comparison_data = load_comparison_data(script_dir)
    withdrawn_controls = comparison_data['withdrawn']
    if withdrawn_controls:
        print(f"Loaded comparison data: {len(withdrawn_controls)} Rev 4-only (withdrawn) controls identified")

    # Separate Rev 4-only controls from Rev 5 controls
    rev5_level_data = {}
    rev4_only_controls = {}

    for level_name, controls in level_data.items():
        rev5_controls = []
        rev4_controls = []

        for ctrl in controls:
            normalized = normalize_control_id(ctrl)
            if normalized in withdrawn_controls:
                rev4_controls.append(normalized)
            else:
                rev5_controls.append(normalized)

        rev5_level_data[level_name] = rev5_controls
        if rev4_controls:
            rev4_only_controls[level_name] = rev4_controls

    # Load Rev 4 reference data if we have Rev 4-only controls
    r4_controls_lookup = {}
    r4_cci_lookup = {}
    if rev4_only_controls:
        r4_controls_path = script_dir / 'r4controls.json'
        r4_cci_path = script_dir / 'rev4cci.json'
        if r4_controls_path.exists():
            print(f"Loading Rev 4 controls for withdrawn controls...")
            r4_controls_lookup = load_controls_data(str(r4_controls_path))
        if r4_cci_path.exists():
            print(f"Loading Rev 4 CCI mappings for withdrawn controls...")
            r4_cci_lookup = load_cci_data(str(r4_cci_path))

    # Create workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Track statistics for summary
    all_stats = {}
    level_names = list(level_data.keys())

    # Create individual level sheets (Rev 5 controls only)
    print("\nGenerating level sheets (Rev 5 controls)...")
    for level_name in level_names:
        controls = rev5_level_data.get(level_name, [])
        rev4_count = len(rev4_only_controls.get(level_name, []))
        rev4_note = f" ({rev4_count} Rev 4-only moved to separate sheet)" if rev4_count > 0 else ""
        print(f"  Creating sheet for {level_name} ({len(controls)} controls){rev4_note}...")
        stats = create_level_sheet(wb, level_name, controls, controls_lookup, cci_lookup)
        all_stats[level_name] = stats

        # Create detailed CCI sheet if requested
        if args.detailed_cci:
            create_cci_detail_sheet(wb, level_name, controls, controls_lookup, cci_lookup)

    # Create Rev 4-only sheet if there are withdrawn controls
    if rev4_only_controls:
        total_rev4 = sum(len(c) for c in rev4_only_controls.values())
        print(f"  Creating Rev 4-only sheet ({total_rev4} withdrawn controls)...")
        create_rev4_only_sheet(wb, rev4_only_controls, r4_controls_lookup, r4_cci_lookup)

    # Create summary sheet
    print("Creating summary sheet with charts...")
    create_summary_sheet(wb, all_stats, level_names)

    # Save workbook
    output_path = Path(output_file) if Path(output_file).is_absolute() else script_dir / output_file
    wb.save(str(output_path))
    print(f"\nWorkbook saved to {output_path}")

    # Print summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)

    # Collect all problematic entries across levels
    all_unknown = []
    all_not_in_ref = []

    for level_name in level_names:
        stats = all_stats.get(level_name, {})
        print(f"\n{level_name}:")
        print(f"  Controls: {stats.get('total_controls', 0)}")
        print(f"  Total CCIs: {stats.get('total_ccis', 0)}")
        families = stats.get('families', {})
        if families:
            print(f"  Families: {', '.join(sorted(families.keys()))}")

        # Collect problematic entries
        unknown = stats.get('unknown_controls', [])
        not_in_ref = stats.get('not_in_reference', [])
        if unknown:
            all_unknown.extend([(level_name, c) for c in unknown])
        if not_in_ref:
            all_not_in_ref.extend([(level_name, c) for c in not_in_ref])

    # Show problematic entries
    if all_unknown:
        print("\n" + "-"*60)
        print(f"WARNING: {len(all_unknown)} entries have 'Unknown' family (invalid format):")
        for level, ctrl in all_unknown[:15]:
            print(f"  [{level[:20]}] {ctrl}")
        if len(all_unknown) > 15:
            print(f"  ... and {len(all_unknown) - 15} more")

    if all_not_in_ref:
        print("\n" + "-"*60)
        print(f"INFO: {len(all_not_in_ref)} controls not found in reference JSON (no name/text):")
        for level, ctrl in all_not_in_ref[:15]:
            print(f"  [{level[:20]}] {ctrl}")
        if len(all_not_in_ref) > 15:
            print(f"  ... and {len(all_not_in_ref) - 15} more")


if __name__ == '__main__':
    main()
